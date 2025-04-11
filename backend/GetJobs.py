import re 
import os 
import pandas as pd
import time
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from dotenv import load_dotenv
from google import genai

WORDS_TO_FILTER = [
    "Home", "Company reviews", "Salary guide", "Employers", "Create your resume",
    "Resume services", "Change country", "Help", "Privacy Centre", "Part-time",
    "Full-time", "Hiring Lab", "Career advice", "Browse jobs", "Browse companies",
    "Salaries", "Indeed Events", "Work at Indeed", "Countries", "About", "Help",
    "ESG at Indeed", "© 2025 Indeed", "Accessibility at Indeed", 
    "Privacy Centre and Ad Choices", "Terms"
]

def configure_driver():
    options = webdriver.ChromeOptions()
    options.add_argument(f'user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
    options.add_argument('--lang=en-CA,en')
    options.add_argument('--accept-language=en-CA,en;q=0.9')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_argument('--disable-notifications')
    options.add_experimental_option('excludeSwitches', ['enable-automation'])
    options.add_experimental_option('useAutomationExtension', False)
    options.add_argument('--window-size=1920,1080')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    
    return driver

def summarize(soup): 
    try:
        load_dotenv() 
        api_key = os.getenv("GEMINI_API_KEY")
        if not api_key:
            return "No summary available - API key missing"
        
        client = genai.Client(api_key=api_key)
        content = soup.get_text()
        
        text_to_summarize = f"Summarize this job posting with key details. User proper punctuation. The following is the content you are to summarize: {content}"
        
        response = client.models.generate_content(
            model="gemini-2.0-flash", 
            contents=[text_to_summarize]
        )
    
        if response.status_code == 200:
            result = response.json()
            if isinstance(result, list) and result[0].get("summary_text"):
                return result[0]["summary_text"]
            else:
                return "No summary available - unexpected response format"
        else:
            error_message = f"API Error: Status {response.status_code}"
            try:
                error_details = response.json()
                error_message += f", Details: {error_details}"
            except:
                error_message += f", Raw response: {response.text}"
            return "No summary available - " + error_message

    except requests.exceptions.RequestException as e:
        return f"No summary available - request error: {str(e)}"
    except Exception as e:
        return f"No summary available - unexpected error: {str(e)}"

def extract_from_summary(summary):
    info = {
        'title': None,
        'company': None,
        'salary': None
    }
    
    salary_pattern = r'(?:From |starting at )?\$[\d,]+(?:\.?\d{2})?(?:\s*-\s*\$[\d,]+(?:\.?\d{2})?)?(?:\s*(?:per hour|an hour))?'
    salary_match = re.search(salary_pattern, summary)
    if salary_match:
        info['salary'] = salary_match.group(0)
    
    title_pattern = r'(?:^|\bis looking for |seeking |hiring )(?:a |an )?([A-Z][A-Za-z\s-]+?)(?=\s+to|\s+who|\s+that|\s+in|\s+at|\s+with|\.|,)'
    title_match = re.search(title_pattern, summary)
    if title_match:
        info['title'] = title_match.group(1).strip()
    
    company_pattern = r'(?:at|for|with|join)\s+([A-Z][A-Za-z\s&.-]+?)(?=\s+to|\s+is|\s+in|\s+seeks|\s+requires|\s+team|\.|,)'
    company_match = re.search(company_pattern, summary)
    if company_match:
        info['company'] = company_match.group(1).strip()
    
    return info

def scrape(url):
    driver = None
    try:
        driver = configure_driver()
        driver.get(url)
        time.sleep(3)
        
        if "indeed.com" not in driver.current_url:
            time.sleep(2)
            
        soup = BeautifulSoup(driver.page_source, "html.parser")
        
        job_title = "Unknown Title"
        title_element = soup.find('h1', {'data-testid': 'jobsearch-JobInfoHeader-title'})
        if title_element:
            job_title = title_element.text.strip()
        
        company = "unknown"
        company_element = soup.find('div', {'data-testid': 'inlineHeader-companyName'})
        if company_element and company_element.find('a'):
            company = company_element.find('a').get_text(strip=True)
        
        location = "Unknown Location"
        location_element = soup.find('div', {'data-testid': 'inlineHeader-companyLocation'})
        if location_element:
            location = location_element.get_text(strip=True)
        
        description = ""
        description_div = soup.find('div', {'id': 'jobDescriptionText'})
        if description_div:
            description = description_div.get_text(strip=True)
        
        job_data = {
            "Job Title": [job_title],
            "Company": [company],
            "Location": [location],
            "Salary": ["unknown"],
            "Description": [description],
            "Responsibilities, Qualifications and/or Benefits": [""],
            "url": [url]
        }
        
        df = pd.DataFrame(job_data)
        save_csv(df, company, job_title)
        return df
        
    except Exception as e:
        raise
    finally:
        if driver:
            driver.quit()

def save_csv(df, company, title): 
    def path_to_file(): 
        home_dir = os.path.expanduser("~")
        file_path = os.path.join(home_dir, "Desktop/jobs.xlsx")
        return file_path

    if not company:
        company = "Unknown_Company"
    if not title:
        title = "Unknown_Title"

    file_path = path_to_file()

    if os.path.exists(file_path):
        existing_df = pd.read_excel(file_path)
        combined_df = pd.concat([existing_df, df])
    else: 
        combined_df = df 

    combined_df.to_excel(file_path, index=False)
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active

        for column_index, column_cells in enumerate(sheet.columns, start=1):
            max_length = 0
            column_letter = get_column_letter(column_index)
            column_name = sheet.cell(row=1, column=column_index).value

            for cell in column_cells:
                if cell.value:
                    cell_value = str(cell.value)
                    max_length = max(max_length, len(cell_value))
                     
                cell.alignment = Alignment(wrap_text=True)

            if column_name == "Description":
                sheet.column_dimensions[column_letter].width = 50
            elif column_name == "Responsibilities, Qualifications and/or Benefits":
                sheet.column_dimensions[column_letter].width = 50
            elif column_name == "Summary":
                sheet.column_dimensions[column_letter].width = 35
            elif column_name == "Job Title":
                sheet.column_dimensions[column_letter].width = 15
            elif column_name in ["Company", "Location", "Salary"]:
                sheet.column_dimensions[column_letter].width = 15
            elif column_name == "url":
                sheet.column_dimensions[column_letter].width = 20
            else:
                adjusted_width = max_length + 2
                sheet.column_dimensions[column_letter].width = adjusted_width
        workbook.save(file_path)
    except Exception as e:
        pass

    
    

import re 
import os 
import pandas as pd

from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# for the summarizer 
import requests
import os
from dotenv import load_dotenv


WORDS_TO_FILTER = [
    "Home",
    "Company reviews",
    "Salary guide",
    "Employers",
    "Create your resume",
    "Resume services",
    "Change country",
    "Help",
    "Privacy Centre",
    "Part-time",
    "Full-time",
    "Hiring Lab",
    "Career advice",
    "Browse jobs",
    "Browse companies",
    "Salaries",
    "Indeed Events",
    "Work at Indeed",
    "Countries",
    "About",
    "Help",
    "ESG at Indeed",
    "© 2025 Indeed",
    "Accessibility at Indeed",
    "Privacy Centre and Ad Choices",
    "Terms"
]

def configure_driver(): 
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")  
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    )
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)
    return driver 

def summarize(soup): 

    load_dotenv() 
    api_key = os.getenv("HUGGINGFACE_API_KEY")
    if not api_key:
        raise ValueError("HUGGINGFACE_API_KEY is not set")
    
    # Use the BART-based summarization model
    API_URL = "https://api-inference.huggingface.co/models/facebook/bart-large-cnn"
    headers = {"Authorization": f"Bearer {api_key}"}

    # kills all html tags
    content = soup.get_text() 

    text_to_summarize = (
        f"Extract and summarize only the following job posting with this structure, seperated with semicolons:\n"
        f"1. Job Title and Company\n"
        f"2. Salary Information (if available)\n"
        f"3. Location\n"
        f"4. Key Requirements (2-3 most important)\n"
        f"5. Main Responsibilities (2-3 most critical)\n"
        f"Content to summarize: {content}"
    )

    # Prepare the correct input format for the BART model
    data = {
        "inputs": text_to_summarize,
        "parameters": {
            "min_length": 50,  # Increased for more detailed summary
            "max_length": 150,  # Increased to accommodate structured format
            "temperature": 0.3,  # Reduced for more focused/consistent output
            "top_p": 0.9,      # Slightly increased for better coherence
        }
    }

    # Send request to Hugging Face API
    try:
        response = requests.post(API_URL, headers=headers, json=data, timeout=300)
        result = response.json() 
        print("Summary:", result[0]["summary_text"])  
        return result[0]["summary_text"]
    
    except Exception as e:
        print("An error occurred:", e)

    return "No summary available"

def extract_from_summary(summary):
    """Extract job details from the summary when primary scraping fails"""
    info = {
        'title': None,
        'company': None,
        'salary': None
    }
    
    # Look for salary patterns including "from" and "per hour" formats
    salary_pattern = r'(?:From |starting at )?\$[\d,]+(?:\.?\d{2})?(?:\s*-\s*\$[\d,]+(?:\.?\d{2})?)?(?:\s*(?:per hour|an hour))?'
    salary_match = re.search(salary_pattern, summary)
    if salary_match:
        info['salary'] = salary_match.group(0)
    
    # Look for job title at the start of summary or after specific phrases
    title_pattern = r'(?:^|\bis looking for |seeking |hiring )(?:a |an )?([A-Z][A-Za-z\s-]+?)(?=\s+to|\s+who|\s+that|\s+in|\s+at|\s+with|\.|,)'
    title_match = re.search(title_pattern, summary)
    if title_match:
        info['title'] = title_match.group(1).strip()
    
    # Look for company name patterns
    company_pattern = r'(?:at|for|with|join)\s+([A-Z][A-Za-z\s&.-]+?)(?=\s+to|\s+is|\s+in|\s+seeks|\s+requires|\s+team|\.|,)'
    company_match = re.search(company_pattern, summary)
    if company_match:
        info['company'] = company_match.group(1).strip()
    
    return info

def scrape(url):
    driver = configure_driver()
    driver.get(url)
    soup = BeautifulSoup(driver.page_source, "html.parser")        
    driver.quit()

    # Get summary first to use as backup
    summary = summarize(soup)
    backup_info = extract_from_summary(summary)

    # Extract job details
    try:
        job_title = soup.find('h1', {'data-testid': 'jobsearch-JobInfoHeader-title'}).text.strip()
        print(f"job title: {job_title}")
    except AttributeError:
        job_title = backup_info['title'] if backup_info['title'] else "Unknown Title"
        print(f"Using backup job title: {job_title}")

    try:   
        company = "unknown"
        company_element = soup.find('div', {'data-testid': 'inlineHeader-companyName'})
        if company_element:
            company = company_element.find('a').get_text(strip=True) 
        if company == "unknown":
            company = backup_info['company'] if backup_info['company'] else "unknown"
        print(f"company: {company}")
    except AttributeError:
        company = backup_info['company'] if backup_info['company'] else "unknown"
        print("Using backup company value")

    try:    
        location_element = soup.find('div', {'data-testid': 'inlineHeader-companyLocation'})
        if location_element:
            location = location_element.get_text(strip=True)
        else:
            location_element = soup.find('div', class_='css-16tkvfy')
            location = location_element.get_text(strip=True) if location_element else "Unknown Location"
    except AttributeError:
        location = "Unknown Location"

    try:
        salary = "unknown"
        salary_element = soup.find('div', {'id': 'salaryInfoAndJobType'})
        if salary_element:
            salary_text = salary_element.find('span').get_text(strip=True)
            pattern = r'\$\S+'
            match = re.search(pattern, salary_text)
            salary = match.group() if match else None
        if salary == "unknown":
            salary = backup_info['salary'] if backup_info['salary'] else "unknown"
    except AttributeError:
        salary = backup_info['salary'] if backup_info['salary'] else "unknown"

    # Extract paragraphs for the job description
    description = ' '.join([p.get_text(strip=True) for p in soup.find_all('p')])

    # Extract any listed benefits or responsibilities
    bullet_points = []
    for li in soup.find_all('li'):
        text = li.get_text(strip=True)
        should_include = True
        for word in WORDS_TO_FILTER:
            if word.lower() in text.lower():
                should_include = False
                break
        if should_include:
            bullet_points.append(f"- {text}")

    list = "\n".join(bullet_points)
    
    # Create and clean a DataFrame for the results
    job_data = {
        "Job Title": [job_title],
        "Company": [company],
        "Location": [location],
        "Salary": [salary],
        "Summary": [summary],
        "Description": [description],
        "Responsibilities, Qualifications and/or Benefits": [list],
        "url": [url]
    }

    df = pd.DataFrame(job_data)
    save_csv(df, company, job_title)
    print(df.to_dict(orient="records"))
    return df

def save_csv(df, company, title): 

    # eventually can get user to prompt where to save to 
    def path_to_file(): 
        home_dir = os.path.expanduser("~")
        file_path = os.path.join(home_dir, "Desktop/jobs.xlsx")
        return file_path

    # Handling no company or title 
    if not company:
        company = "Unknown_Company"
    if not title:
        title = "Unknown_Title"

    file_path = path_to_file() # gets the file path 

    if os.path.exists(file_path):
        existing_df = pd.read_excel(file_path)
        combined_df = pd.concat([existing_df, df])
    else: 
        combined_df = df 

    #2 write the combined dataframe back to excel 
    combined_df.to_excel(file_path, index=False)
    # Save the DataFrame to an Excel file
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active

        for column_index, column_cells in enumerate(sheet.columns, start=1):
            max_length = 0
            column_letter = get_column_letter(column_index)
            column_name = sheet.cell(row=1, column=column_index).value  # Get the column name

            for cell in column_cells:
                if cell.value:
                    cell_value = str(cell.value)
                    max_length = max(max_length, len(cell_value))
                     
                # Enable text wrapping for all cells
                cell.alignment = Alignment(wrap_text=True)

            # Set specific widths for each column type
            if column_name == "Description":
                sheet.column_dimensions[column_letter].width = 50
            elif column_name == "Responsibilities, Qualifications and/or Benefits":
                sheet.column_dimensions[column_letter].width = 50
            elif column_name == "Summary":
                sheet.column_dimensions[column_letter].width = 35  # Reduced from default
            elif column_name == "Job Title":
                sheet.column_dimensions[column_letter].width = 15
            elif column_name in ["Company", "Location", "Salary"]:
                sheet.column_dimensions[column_letter].width = 15
            elif column_name == "url":
                sheet.column_dimensions[column_letter].width = 20
            else:
                adjusted_width = max_length + 2  # add padding
                sheet.column_dimensions[column_letter].width = adjusted_width
        # Save the formatted Excel file
        workbook.save(file_path)
        print(f"File saved successfully with formatting to {file_path}")
    except Exception as e:
        print(f"Error saving file: {e}")

    
    
